import json
import ntpath
import webbrowser
from time import gmtime, strftime

import requests
from PyQt5 import QtWidgets, uic, QtGui, QtCore
import os, sys, datetime, pymysql, threading, pandas as pd

from PyQt5.QtGui import QPainter, QPalette
from PyQt5.QtWidgets import QHeaderView, QGraphicsDropShadowEffect
from openpyxl.styles import Border, Side
from plyer import notification
from openpyxl import load_workbook
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
# import pyrebase
from pyrebase import pyrebase

radius = 40.0


today = datetime.datetime.today().strftime('%Y-%m-%d')


def notif(self = None, title = '', msg = ''):
    try:
        notification.notify(
            title=title,
            message=msg,
            app_icon=os.path.join(os.getcwd(), 'src/img/it.ico'),
            timeout=5,
        )
    except:
        QtWidgets.QMessageBox.about(self, title, msg)


def con():
    return pymysql.connect(host='10.73.100.101', user='VP', database='vansplanning', password='1234@@it.', port=3306)


def preparingDB():
    cnx = con()
    cur = cnx.cursor()
    cur.execute('''Create table if not exists users (id  INT AUTO_INCREMENT, firstName Varchar(50), LastName Varchar(50), username Varchar(50), pass Varchar(50), role INT , PRIMARY KEY (id));''')
    cur.execute('''create table if not exists grps (id  INT AUTO_INCREMENT, name Varchar(50), shift Varchar(20) , PRIMARY KEY (id));''')
    cnx.commit()
    cur.execute('''create table if not exists agents (id  INT AUTO_INCREMENT, matrr varchar(50) not null, firstName Varchar(50), LastName Varchar(50), CIN varchar(50), address Varchar(100), grp int, zone Varchar(45), street  Varchar(45), pic text, PRIMARY KEY (id), foreign key(grp) references grps(id));''')
    cur.execute('''create table if not exists vans (id  INT AUTO_INCREMENT, matr Varchar(50), driver varchar(50), max_places int , PRIMARY KEY (id))''')
    cur.execute('''create table if not exists vans(id  INT AUTO_INCREMENT, matricule varchar(50), max_places int , PRIMARY KEY (id))''')
    cur.execute('''create table if not exists drivers(id  INT AUTO_INCREMENT, firstName Varchar(50), LastName varchar(50), username Varchar(45), pass Varchar(45), PRIMARY KEY (id))''')
    cnx.commit()

    cur.execute('create table if not exists trips(id  INT AUTO_INCREMENT, van int, driver int, datetime Varchar(50), ttype Varchar(5), starttime varchar(45), startloc varchar(45), stoptime varchar(45), stoploc varchar(45), counterkm varchar(45), tracking LONGTEXT, foreign key(van) references vans(id), foreign key(driver) references drivers(id) , PRIMARY KEY (id))')

    cnx.commit()
    cur.execute('''create table if not exists trips_history(id  INT AUTO_INCREMENT,trip int, agent int, presence varchar(50), picktime varchar(50), pickloc varchar(50), droptime varchar(50), droploc varchar(50), foreign key(trip) references trips(id), foreign key(agent) references agents(id), PRIMARY KEY (id))''')
    cnx.commit()
    cur.execute("CREATE TABLE if not exists logsfile (id INT AUTO_INCREMENT,user VARCHAR(45) ,query VARCHAR(500) ,status VARCHAR(50), descr VARCHAR(500) ,PRIMARY KEY (id))")
    cnx.commit()
    # cur.execute()
    cnx.close()
    #print('tables created ')


# preparingDB()

print('test')

# def createPlannings():
#     today = datetime.date.today()
#     day = datetime.datetime.today().strftime('%A')
#     #print(day)
#     if day.lower() == "friday":
#         tomorrow = datetime.date.today() + datetime.timedelta(days=3)
#     else:
#         tomorrow = datetime.date.today() + datetime.timedelta(days=1)
#
#     # tomorrow = today #TODO Just for test
#     doneDates = []
#     for date in [today, tomorrow]:
#         #print(f'date ==> {date}')
#         cnx = con()
#         cur = cnx.cursor()
#         cur.execute('select shift from grps')
#         data = []
#         for r in cur.fetchall():
#             for c in r:
#                 for cc in str(c).split('-'):
#                     data.append(int(cc))
#         dt = [i for i in set(sorted(data))]
#         data = []
#         for i in dt:
#             if i < 10:
#                 data.append(f'0{i}')
#             else:
#                 data.append(str(i))
#         #print(data)
#         trips = {}
#         total = 0
#         ids = []
#         for h in data:
#             query = ''
#             if int(h) < 15:
#                 query = f'select a.id from agents a inner join grps g on a.grp = g.id where g.shift like "{h}-%"'
#             else:
#                 query = f'select a.id from agents a inner join grps g on a.grp = g.id where g.shift like "%-{h}"'
#
#             cur.execute(query)
#             result = cur.fetchall()
#             # ##print(result)
#
#             count = len(result)
#
#             trips[h] = [i[0] for i in result]
#             total += count
#
#         ##print(trips)
#         ##print(total)
#         maxForVan = 18
#         v1 = [8, 13, 17, 22]
#         v2 = [9, 14, 18, 23]
#         for time, agents in trips.items():
#
#             vansNeeded = 0
#             if agents:
#                 if len(agents) > maxForVan:
#                     vansNeeded = 2
#                 else:
#                     vansNeeded = 1
#
#
#                 if vansNeeded == 1:
#                     cur.execute(
#                         f''' select count(id) from trips where datetime like "{date} {time}:00:00" and van = 1''')
#                     if not cur.fetchone()[0]:
#                         cur.execute(
#                             f'select g.shift from agents a inner join grps g on g.id = a.grp where a.id = {agents[0]}')
#
#                         cur.execute(
#                             f'''insert into trips(van, driver, datetime, ttype) values (1, 1, "{date} {time}:00:00", '{"IN" if [i for i in str(cur.fetchone()[0]).split('-')].index(time) == 0 else "OUT"}');''')
#                         cnx.commit()
#                         for agent in agents:
#
#
#                             cur.execute(f'''select id from trips where datetime like "{date} {time}:00:00"''')
#                             cur.execute(
#                                 f'''insert into trips_history (trip, agent, presence) values ({int(cur.fetchone()[0])}, {agent}, 0)''')
#                             cnx.commit()
#                 else:
#                     grp1 = [i for i in agents[:len(agents) // 2]]
#                     grp2 = [i for i in agents[len(agents) // 2:]]
#                     cur.execute(
#                         f'''select count(id) from trips where datetime like "{date} {time}:00:00" and driver = 1''')
#                     if not cur.fetchone()[0]:
#                         cur.execute(
#                             f'''insert into trips(van, driver, datetime) values (1, 1, "{date} {time}:00:00");''')
#
#                     cur.execute(f'''select id from trips where datetime like "{date} {time}:00:00"''')
#
#                     trip1, trip2 = [int(i[0]) for i in cur.fetchall()]
#                     for agent in grp1:
#                         cur.execute(
#                             f'''select count(id) from trips_history where trip = {trip1} and agent = {agent}''')
#                         if not cur.fetchone()[0]:
#                             cur.execute(
#                                 f'''insert into trips_history (trip, agent, presence) values ({trip1}, {agent}, 0)''')
#                             cnx.commit()
#
#                     cur.execute(
#                         f'''select count(id) from trips where datetime like "{date} {time}:00:00" and driver = 1''')
#                     if not cur.fetchone()[0]:
#                         cur.execute(
#                             f'''insert into trips(van, driver, datetime) values (2, 2, "{date} {time}:00:00");''')
#                     cnx.commit()
#                     for agent in grp2:
#                         cur.execute(
#                             f'''select count(id) from trips_history where trip = {trip2} and agent = {agent}''')
#                         if not cur.fetchone()[0]:
#                             cur.execute(
#                                 f'''insert into trips_history (trip, agent, presence) values ({trip2}, {agent}, 0)''')
#                             cnx.commit()
#         doneDates.append(f'{date}')
#     cnx.close()
#
#     notif(title="Saccom IT Departement : ", msg=f'{len(doneDates)} Plannings has been created : {doneDates} ')


def create_trip(date, type, driver, van):
    cnx = con()
    cur = cnx.cursor()
    cur.execute(f'''select id from vans where matr like "{van}"''')
    van = int(cur.fetchone()[0])
    cur.execute(f'''select id from drivers where firstName like "{driver.split(" ")[0]}" and LastName like "{driver.split(" ")[1]}"''')
    driver = int(cur.fetchone()[0])

    tday, trips_time, trip_type = date.split(' ')[0], int(date.split(' ')[1].split(':')[0]), type
    cur.execute(f'''select count(id) from trips where datetime like "{date}";''')
    if not cur.fetchone()[0]:
        cur.execute(f'''select max_places from vans where id = {van}''')
        max_places = int(cur.fetchone()[0])
        beforHour = trips_time - 1
        afterHour = trips_time + 1
        cur.execute(f'''select count(id) from trips 
        where (datetime like "{tday} {beforHour if beforHour > 9 else f"0{beforHour}"}:00:00" 
        OR datetime like "{tday} {afterHour if afterHour > 9 else f"0{afterHour}"}:00:00") 
        and driver = {driver} and van = {van};''')
        if not cur.fetchone()[0]:

            if 6 > trips_time > 20:
                cur.execute('''select zone from agents''')
                zones = set(i[0] for i in cur.fetchall())
                zone1 = []
                zone2 = []
                for z in zones:
                    if z in ['Sidi Yahya', 'Sidi Ziane']:
                        zone1.append(z)
                    else:
                        zone2.append(z)

                cur.execute(
                    f'''select id from agents where zone in ({[i for i in zone1]})'''.replace('[', '').replace(']', ''))

                agentZone1 = [i[0] for i in cur.fetchall()]

                cur.execute(
                    f'''select id from agents where zone in ({[i for i in zone2]})'''.replace('[', '').replace(']', ''))

                agentZone2 = [i[0] for i in cur.fetchall()]

                for i in range(1, 3):
                    cur.execute(
                        f'''insert into trips (van, driver, datetime, ttype)values ({i}, {i}, "{tday} {trips_time if trips_time > 9 else f"0{trips_time}"}:00:00", "{trip_type}")''')
                    cnx.commit()
                    cur.execute(f'''select id from trips where datetime like "{date}" and van = {i} and diver = {i}''')
                    trip_id = cur.fetchone()[0]
                    if i == 1:
                        cur.execute(
                            f'''insert into trips_histoy(trip, agent) values {[f"({trip_id}, {agent})" for agent in agentZone1]}'''.replace(
                                '[', '').replace(']', '').replace("'", ""))

                    if i == 2:
                        cur.execute(
                            f'''insert into trips_histoy(trip, agent) values {[f"({trip_id}, {agent})" for agent in agentZone2]}'''.replace(
                                '[', '').replace(']', '').replace("'", ""))


                    cnx.commit()
                    Trip_View(role=0, id=(trip_id,))

            else:
                cur.execute(
                    f'''insert into trips (van, driver, datetime, ttype)values ({van}, {driver}, "{tday} {trips_time if trips_time > 9 else f"0{trips_time}"}:00:00", "{trip_type}")''')
                cnx.commit()
                cur.execute(f'''select id from trips where datetime like "{date}"''')
                trip_id = cur.fetchone()[0]
                cur.execute(f'''select a.id from agents a inner join grps g on a.grp = g.id inner join trips_history th on th.agent = a.id inner join trips t on th.trip = t.id
                where g.shift like "{f"{trips_time if trips_time > 9 else f'0{trips_time}'}-%" if trip_type == 'IN' else f"%-{trips_time if trips_time > 9 else f'0{trips_time}'}"}" 
                and t.datetime != "{date}" order by a.zone''')
                agents = [i[0] for i in set(cur.fetchall())]
                print(agents)
                if agents:
                    if len(agents) > max_places:

                        grp1 = [i for i in agents[:len(agents) // 2]]
                        for a in grp1:
                            # query = f'''insert into trips_history(trip, agent) values {[f"({trip_id}, {agent})" for agent in grp1]}'''.replace('[', '').replace(']', '').replace("'", "")
                            query = f'''insert into trips_history(trip, agent) values({trip_id}, {a}) ;'''
                            print(query)
                            cur.execute(query)
                            cnx.commit()
                        Trip_View(role=0, id=(trip_id,)).show()
                        cur.execute(
                            f'''select d.id from drivers d inner join trips t on t.driver = d.id  where d.id != (select driver from trips where id = {trip_id})''')
                        driver2 = cur.fetchone()[0]
                        cur.execute(
                            f'''select v.id from vans v inner join trips t on t.van = v.id  where v.id != (select van from trips where id = {trip_id})''')
                        van2 = cur.fetchone()[0]
                        create_trip(date=date, driver=driver2, van=van2, type=trip_type)
                    else:
                        for a in agents:
                            # query = f'''insert into trips_history(trip, agent) values {[f"({trip_id}, {agent})" for agent in agents]}'''.replace('[', '').replace(']', '').replace("'", "")
                            query = f'''insert into trips_history(trip, agent) values({trip_id}, {a}) ;'''
                            print(query)
                            cur.execute(query)
                            cnx.commit()
                        Trip_View(role=0, id=(trip_id,)).show()






        else:
            print('cant create this trip right now')

    else:
        print('the trip already exists ...')

    cnx.close()


config = {
"apiKey": "AIzaSyAIV0qDv7WI3kU_krALZEc-PlTCkBRMp9g",
"authDomain": "saccomxd-stm-yassine-baghdadi.firebaseapp.com",
"databaseURL": "https://saccomxd-stm-yassine-baghdadi-default-rtdb.firebaseio.com",
"projectId": "saccomxd-stm-yassine-baghdadi",
"storageBucket": "saccomxd-stm-yassine-baghdadi.appspot.com",
"serviceAccount": os.path.join(os.getcwd(), "src", "key.json")
}
token = "e230e05d-8b62-4d8b-8b17-56d929d19d54"

def uploadPicToFireBase(img : str) -> str:
    if os.path.isfile(img):
        print(f"Local path ==> {img}")
        firebase_storage = pyrebase.initialize_app(config)
        storage = firebase_storage.storage()
        storage.child(ntpath.basename(img)).put(img)
        return storage.child(ntpath.basename(img)).get_url(token)


cred = credentials.Certificate("src/key.json")
firebase_admin.initialize_app(cred, {
        'databaseURL': 'https://saccomxd-stm-yassine-baghdadi-default-rtdb.firebaseio.com/',

    })
def syncFirebase():
    print('synchronizing with Firebase ...')
    # Fetch the service account key JSON file contents

    # Initialize the app with a service account, granting admin privileges


    # users = db.reference('users')
    # users = ref.child('users')
    # print(users.child('YassineBaghdadi').get())
    cnx = con()
    cur = cnx.cursor()
    # cur.execute('select concat(firstName, LastName) as fullName , username, pass from users;')
    # dt = cur.fetchall()
    # for d in dt:
    #     if not users.child(f'{d[0]}').get():
    #         users.child(f'{d[0]}').set({
    #             'username': d[1],
    #             'pass': d[2],
    #         }
    #         )

    drvrs = db.reference('drivers')
    cur.execute('select username, pass from drivers;')
    users = cur.fetchall()
    for d in users:
        if not drvrs.child(f'{d[0]}').get():
            drvrs.child(f'{d[0]}').set({'pass': d[1],})
        else:
            drvrs.child(d[0]).child("pass").set(d[1])




    cur.execute(f'''select t.id, v.matr, d.username , t.datetime, (select count(id) from trips_history where trip = t.id) as Agent_numbers, t.ttype
                        from trips t inner join vans v on t.van = v.id inner join drivers d on t.driver = d.id where date(t.datetime) >= "{today} 00:00:00" order by t.datetime;''')

    # cur.execute(f'''select t.id, v.matr, d.username , t.datetime, (select count(id) from trips_history where trip = t.id) as Agent_numbers, t.ttype
    #                     from trips t inner join vans v on t.van = v.id inner join drivers d on t.driver = d.id order by t.datetime;''')

    trips = cur.fetchall()

    for trip in trips:
        trp = drvrs.child(f'{trip[2]}').child('trips').child(f'{trip[3]}')
        cur.execute(f'''select a.id, concat(a.firstName, " ", a.LastName) as fullName,
                                             g.name, a.pic from trips_history th inner join agents a on th.agent = a.id inner join grps g on a.grp = g.id
                                             where th.trip = {int(trip[0])};''')

        agents = [{"name" : f"{i[0]} - {i[1]}", "phone" : "0630504606", "pic" : i[3], "grp" : i[2], "droptime":"", "droploc":"", "picktime":"", "pickloc":"", "presence":""} for i in cur.fetchall()]


        if not trp.get():
            trp.set({
                            'id' : trip[0],
                            'agents' : agents,

                        })
            print(f'Agents : {agents}')
        else:
            print(f'{trip[0]} alredy added')



    # cur.execute(f'''select t.id, v.matr, d.username , t.datetime, (select count(id) from trips_history where trip = t.id) as Agent_numbers, t.ttype
    #                     from trips t inner join vans v on t.van = v.id inner join drivers d on t.driver = d.id where date(t.datetime) < "{today} 00:00:00" order by t.datetime;''')
    cur.execute(f'''select datetime 
                        from trips where date(datetime) < "{today} 00:00:00" order by datetime;''')

    toDelete = [i[0] for i in cur.fetchall()]

    drivers = [k for k, v in dict(drvrs.get()).items()]
    for dr in drivers:
        if drvrs.child(dr).child('trips').get() :
            for tt in dict(drvrs.child(dr).child('trips').get()).keys():
                if tt in toDelete:
                    tr = drvrs.child(dr).child('trips').child(tt)
                    tripID = tr.child("id").get()
                    print(f"tripID = {tripID}")


                    trck = tr.child("tracking").get()

                    trajet = {}
                    if trck:
                        for key, value in trck.items():
                            if value not in trajet.values() and value != '0':
                                trajet[key] = value


                    cur.execute(f'''update trips set
                                    starttime = "{tr.child("starttime").get()}",
                                    startloc = "{tr.child("startloc").get()}",
                                    stoptime = "{tr.child("stoptime").get()}",
                                    stoploc = "{tr.child("stoploc").get()}",
                                    counterkm = "{int(str(tr.child("stopcounter").get())) - int(str(tr.child("startcounter").get())) if tr.child("stopcounter").get() and  tr.child("startcounter").get() else 0} ",
                                    tracking = "{trajet}"
                                    where id = {tripID}
                                    ''')
                    cnx.commit()
                    print(tr.child("agents").get())
                    old_agents = list(tr.child("agents").get())
                    if old_agents:
                        for ag in old_agents:
                            try:
                                agent_id = int(str(ag["name"]).split(' - ')[0])
                                cur.execute(
                                    f'''update trips_history set picktime = "{ag['picktime']}", pickloc = "{ag["pickloc"]}", droptime = "{ag["droptime"]}", droploc = "{ag["droploc"]}", presence = "{ag["presence"]}"
                                where trip = {tripID} and agent = {agent_id}''')
                                cnx.commit()
                            except Exception as e:
                                # cur.execute(f"insert into logs(user, query, status, desc)value("", "", "", "")")
                                # cnx.commit()
                                print(e)
                    tr.delete()





    print(trips)
    print('trips synchronized with FireBase successfully.')
    notification.notify(
        title="STM",
        message='trips synchronized with FireBase successfully.',
        app_icon=os.path.join(os.getcwd(), 'src/img/it.ico'),
        timeout=5,
    )

    cnx.close()


class Splash(QtWidgets.QDialog):
    def __init__(self):
        super(Splash, self).__init__()
        QtWidgets.QWidget.__init__(self)
        uic.loadUi(os.path.join(os.path.dirname(__file__), "ui/splash.ui"), self)


        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.setWindowTitle('welcome back')
        self.timer = QtCore.QBasicTimer()
        self.step = 0
        priparingDB = threading.Thread(target=preparingDB)
        priparingDB.start()

        creatingPlanningThread = threading.Thread(target=syncFirebase)
        creatingPlanningThread.start()

        self.prog()

        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

    def paintEvent(self, event):
        # get current window size
        opt = QtWidgets.QStyleOption()
        opt.initFrom(self)
        rect = opt.rect

        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        p.setBrush(opt.palette.brush(QPalette.Window))
        p.setPen(QtCore.Qt.NoPen)
        p.drawRoundedRect(rect, 40, 40)
        p.end()

    def isConnected(self):
        conn = con()
        #todo prepaire the error interface
        if not conn:
            ##print('can\'t connect to the database ...')
            exit()
        conn.close()
        return True




    def prog(self):
        if self.timer.isActive():
            self.timer.stop()
        else:
            self.timer.start(100, self)

    def timerEvent(self, event):
        if self.step >= 100 :
            self.timer.stop()

            self.logi = LogIn()
            self.logi.show()
            self.close()
            return
        self.step += 4
        self.progressBar.setValue(self.step)


class LogIn(QtWidgets.QWidget):
    def __init__(self):
        super(LogIn, self).__init__()
        QtWidgets.QWidget.__init__(self)
        uic.loadUi(os.path.join(os.path.dirname(__file__), "ui/logIn.ui"), self)
        # path.addRoundedRect(QtCore.QRectF(self.rect()), radius, radius)
        # self.setWindowFlag(QtCore.Qt.WindowCloseButtonHint, False)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)

        self.cnx.setEnabled(False)


        # palette = QPalette()
        # palette.setBrush(QPalette.Background, QBrush(QImage("img/pack.png")))
        # self.setPalette(palette)
        self.label.setPixmap(QtGui.QPixmap('src/img/login.png'))
        self.label.setScaledContents(True)
        self.setWindowTitle('Transporting Planning ')
        self.username_in.setPlaceholderText('User Name')
        self.passwrd_in.setPlaceholderText('Password')
        self.cancel.clicked.connect(lambda : os._exit(0))

        self.username_in.textChanged.connect(self.typing)
        self.passwrd_in.textChanged.connect(self.typing1)
        self.passwrd_in.returnPressed.connect(self.login)
        self.username_in.returnPressed.connect(self.login)
        self.cnx.clicked.connect(self.login)
        self.label_2.setPixmap(QtGui.QPixmap('src/img/logo.png'))
        self.label_2.setScaledContents(True)
        self.offset = None
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

    def paintEvent(self, event):
        # get current window size
        opt = QtWidgets.QStyleOption()
        opt.initFrom(self)
        rect = opt.rect

        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        p.setBrush(opt.palette.brush(QPalette.Window))
        p.setPen(QtCore.Qt.NoPen)
        p.drawRoundedRect(rect, 40, 40)
        p.end()

    def mousePressEvent(self, event):
            if event.button() == QtCore.Qt.LeftButton:
                self.offset = event.pos()
            else:
                super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
            if self.offset is not None and event.buttons() == QtCore.Qt.LeftButton:
                self.move(self.pos() + event.pos() - self.offset)
            else:
                super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
            self.offset = None
            super().mouseReleaseEvent(event)




    def typing(self):
        if len(self.username_in.text()) > 2:
            self.username_in.setStyleSheet('border : 2px solid white')
        else:
            self.username_in.setStyleSheet('border : 2px solid red')

        if len(self.username_in.text()) > 2 and len(self.passwrd_in.text()) > 2:
            self.cnx.setEnabled(True)
        else:
            self.cnx.setEnabled(False)



    def typing1(self):
        if len(self.passwrd_in.text()) > 2:
            self.passwrd_in.setStyleSheet('border : 2px solid white')
            self.cnx.setEnabled(True)
        else:
            self.passwrd_in.setStyleSheet('border : 2px solid red')
            self.cnx.setEnabled(False)

        if len(self.username_in.text()) > 2 and len(self.passwrd_in.text()) > 2:
            self.cnx.setEnabled(True)
        else:
            self.cnx.setEnabled(False)


    def login(self):
        username = self.username_in.text()
        passwrd = self.passwrd_in.text()

        ##print(f'UserName : {username}\nPassword : {passwrd}')
        conn = con()
        cur = conn.cursor()
        try:
            cur.execute(f'''SELECT role FROM users Where username like "{username}" and pass like "{passwrd}";''')
            role = cur.fetchone()
            if role is not None:
                self.main = Main(role=int(role[0]))
                self.main.show()
                self.close()
            else:
                self.passwrd_in.setStyleSheet('border : 2px solid red')
                self.username_in.setStyleSheet('border : 2px solid red')
                self.passwrd_in.clear()
                # python -m pip install plyer
                from plyer import notification
                notif(self, title='Transport Planning ', msg='The Username Or The Passsword are wrong')

        except Exception as e:
            print(e)

        conn.close()


class Main(QtWidgets.QWidget):
    def __init__(self, role):
        super(Main, self).__init__()
        QtWidgets.QWidget.__init__(self)
        uic.loadUi(os.path.join(os.path.dirname(__file__), "ui/main.ui"), self)

        self.role = role

        if self.role == 0:
            print('The Admin')

        self.setWindowTitle('Home Page')
        width = 120
        height = 67
        # self.emps.clicked.connect(self.openAgentsUI)
        self.emps.installEventFilter(self)
        self.emps_icon.setPixmap(QtGui.QPixmap('src/img/emps.png'))

        self.vans_icon.installEventFilter(self)
        self.vans_icon.setPixmap(QtGui.QPixmap('src/img/vans.png'))

        self.comps_icon.installEventFilter(self)
        self.comps_icon.setPixmap(QtGui.QPixmap('src/img/comps.png'))

        self.plan_icon.installEventFilter(self)
        self.plan_icon.setPixmap(QtGui.QPixmap('src/img/planning.png'))
        self.emps_icon.setScaledContents(True)

        self.about.installEventFilter(self)
        self.about.setPixmap(QtGui.QPixmap('src/img/about.png'))
        # tip = '''STM 0.1 SACCOM TRANSPORT MANAGEMENT\nCreated by Yassine Baghdadi\nwith help from Anass Kada\nbig thanks to Saccom IT team'''
        # self.about.setToolTip(tip)
        self.about.setScaledContents(True)
        ##print(self.plan_icon.width(), self.plan_icon.height())
        # self.emps.setFixedSize()


        self.planning.installEventFilter(self)
        head = ['Trip ID','Van Matricule','Driver name ','Trip time','Agents numbers', 'Trip Type']
        # self.trips.setHeaderLabels(head)

        self.tripsList = {}
        self.tripsDetails = {}
        self.getCurrentPlanning()
        # self.trips.itemSelectionChanged.connect(lambda : print(f'Select ==> {self.trips.selectedItems()[0].text(0)}'))
        # self.trips.itemDoubleClicked.connect(self.trip_View)
        # self.trips.itemDoubleClicked.connect(lambda : [print(ind.data()) for ind in self.trips.selectedItems()])
        self.comps.installEventFilter(self)
        self.vans.installEventFilter(self)
        STYLESHEET = '''QAbstractItemView QHeaderView {
                show-decoration-selected: 0;
                
            }
            QAbstractItemView::section QHeaderView::section {
                show-decoration-selected: 0;
                background: transparent;
            }
            '''
        # self.trips.setStyleSheet(STYLESHEET)

        self.cls.setPixmap(QtGui.QPixmap('src/img/cls.png'))
        self.cls.setScaledContents(True)
        self.noContentIcon.setPixmap(QtGui.QPixmap('src/img/noContent.png'))
        self.noContentIcon.setScaledContents(True)
        self.cls.installEventFilter(self)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.offset = None
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        # print(self.tripsInfos)
        self.trips_list.itemSelectionChanged.connect(self.trips_selectionChanged)
        self.trips_selectionChanged()
        self.view_btn.clicked.connect(self.trip_View)




    def paintEvent(self, event):
        # get current window size
        opt = QtWidgets.QStyleOption()
        opt.initFrom(self)
        rect = opt.rect

        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        p.setBrush(opt.palette.brush(QPalette.Window))
        p.setPen(QtCore.Qt.NoPen)
        p.drawRoundedRect(rect, 40, 40)
        p.end()

    def mousePressEvent(self, event):
            if event.button() == QtCore.Qt.LeftButton:
                self.offset = event.pos()
            else:
                super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
            if self.offset is not None and event.buttons() == QtCore.Qt.LeftButton:
                self.move(self.pos() + event.pos() - self.offset)
            else:
                super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
            self.offset = None
            super().mouseReleaseEvent(event)

    def eventFilter(self, s, e):
        if e.type() == QtCore.QEvent.MouseButtonPress:
            if s is self.emps:
                self.openAgentsUI()

            if s is self.vans:
                self.openVans()

            if s is self.comps:
                self.openComps()

            if s is self.planning:
                self.openTrips()

            if s is self.about:
                webbrowser.open("https://yassinebaghdadi.github.io/")


        # if e.type() == QtCore.QEvent.Enter:
        #     if s  in (self.emps, self.vans, self.comps, self.planning):
        #         self.inhover(s=s)
        #
        # if e.type() == QtCore.QEvent.Leave:
        #     if s  in (self.emps, self.vans, self.comps, self.planning):
        #         self.outhover(s=s)

        if s is self.cls:
            if e.type() == QtCore.QEvent.Enter:
                self.cls.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                ##print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.cls.setStyleSheet('border-color: 0px ;')
                ##print('mouse out')
            elif e.type() == QtCore.QEvent.MouseButtonPress:
                os._exit(0)



        return super(Main, self).eventFilter(s, e)

    def inhover(self, s):
        s.setStyleSheet("border: 1px solid black; border-radius: 25px;")

    def outhover(self, s):
        s.setStyleSheet('border-color: 0px ;')


    def openVans(self):
        self.vansUI = Vans(role=self.role)
        self.vansUI.show()
        self.close()

    def openComps(self):

        self.compUI = Compaigns(role=self.role)
        self.compUI.show()
        self.close()

    def trip_View(self):

        r =  [self.trips_list.currentIndex().siblingAtColumn(i).data() for i in range(self.trips_list.columnCount())]
        tripID = int(r[0]) if r else None
        if tripID:
            ##print(f'data ==> {data}')
            cnx = con()
            cur = cnx.cursor()
            cur.execute(f'select datetime from trips where id = {tripID} ;')
            dd = cur.fetchone()[0]
            ##print(f'thr date : {dd}')
            trDate = datetime.datetime.strptime(f"{str(dd).split(' ')[0]} {str(dd).split(' ')[1].split(':')[0]}", '%Y-%m-%d %H')
            now = datetime.datetime.strptime(datetime.datetime.today().strftime('%Y-%m-%d %H'), '%Y-%m-%d %H')

            # ##print(f'{"*"*100}\nnow ({now}) > trDate ({trDate}) = {now > trDate}\n{"*"*100}')

            # if int(strftime("%H", gmtime())) >= int(str(dd).split(' ')[1].split(':')[0]) and int(str(dd).split(' ')[0].split('-')[2]) < int(strftime("%d", gmtime())) and int(str(dd).split(' ')[0].split('-')[1]) <= int(strftime("%m", gmtime())) :
            if now > trDate:
                ##print('the combo should desibled ')
                self.TV = Trip_View(role=self.role, id=tripID, desibleedit=True)
                notif(self, title="Error", msg='You can\'t Modify this Trip .')
            else:
                self.TV = Trip_View(role=self.role, id=tripID)
            self.TV.show()
            self.close()




    def openTrips(self):
        self.tri = Trips(role=self.role)
        self.tri.show()
        self.close()

    def openAgentsUI(self):
        self.agents = Agents(self.role)
        self.agents.show()
        self.close()


    def trips_selectionChanged(self):
        self.trip_info.setHorizontalHeaderLabels(
            [i for i in "Matriculation.Full Name.Group.Address.Zone.Shift ".split('.')])
        print(self.tripsDetails)
        selectedRow = [self.trips_list.currentIndex().siblingAtColumn(i).data() for i in range(self.trips_list.columnCount())]
        tripID = selectedRow[0]
        print(f"selected id ==> {tripID}")
        if not tripID:
            self.noContentFrame.setFixedHeight(393)
            self.trip_info.setFixedHeight(0)
            self.view_btn.setEnabled(False)
        else:

            self.noContentFrame.setFixedHeight(0)
            self.trip_info.setFixedHeight(393)
            self.view_btn.setEnabled(True)
            self.trip_info.clear()

            # self.trip_info.setColumnCount(len(self.tripsList[tripID]))
            print(self.tripsDetails[int(tripID)])
            agents = self.tripsDetails[int(tripID)]
            self.trip_info.setRowCount(len(agents))
            self.trip_info.setColumnCount(len(agents[0]))
            self.trip_info.setHorizontalHeaderLabels(
                [i for i in "Matriculation.Full Name.Group.Address.Zone.Shift ".split('.')])
            for i in range(self.trip_info.columnCount()):
                if i == 3:
                    self.trip_info.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeToContents)
                else:
                    self.trip_info.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)
                # self.trips_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)


            for idx, agent in enumerate(agents):
                for cidx, v in enumerate(agent):
                    self.trip_info.setItem(idx, cidx, QtWidgets.QTableWidgetItem(str(v)))






    def getCurrentPlanning(self):
        tread = threading.Thread(target=syncFirebase)
        tread.start()
        # self.trips_list.setRowCount(0)
        # self.trips.clear()
        currentHour = strftime("%H", gmtime())
        currentMinute = strftime("%M", gmtime())
        ##print(f'Current Minute : {currentMinute}')

        cnx = con()
        cur = cnx.cursor()

        cur.execute(f'''select id, datetime
                    from trips where date(datetime) >= "{today} 00:00:00" order by datetime;''')
        data = cur.fetchall()


        ind = 0
        data = [[c for c in r] for r in data]
        for trip in data:
            self.tripsList[trip[0]] = trip[1]
            cur.execute(f'''select a.matrr,  concat(a.firstName, " ", a.LastName), g.name, concat(a.address, " ", a.street), a.zone, g.shift
                                from trips_history th inner join agents a on th.agent = a.id
                                inner join grps g on a.grp = g.id where th.trip = {trip[0]}''')
            self.tripsDetails[trip[0]] = [[i for i in agent] for agent in cur.fetchall()]



        self.trips_list.setColumnCount(2)

        for i in  self.tripsList.keys():
            rowPosition = self.trips_list.rowCount()
            self.trips_list.insertRow(rowPosition)
            self.trips_list.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(f"{i}"))
            self.trips_list.setItem(rowPosition, 1, QtWidgets.QTableWidgetItem(f"{self.tripsList[i]}"))


        self.trips_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.trips_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)

        # for i in data:
        #     # i.append('IN' if str(i[3]).split(' ')[1].split(':')[0] in "08 09 13 14".split() else "Out")
        #     item = QtWidgets.QTreeWidgetItem([str(u) for u in i])
        #     cur.execute(f'''select a.matrr, CONCAT(a.firstName, " ", a.LastName) as agentName, th.picktime, th.presence, g.name  from trips t inner join trips_history th on th.trip = t.id inner join agents a on th.agent = a.id inner join grps g on a.grp = g.id where t.id = {i[0]};''')
        #     agentsForTrip = [i for i in cur.fetchall()]
        #     if agentsForTrip:
        #         ch1 = QtWidgets.QTreeWidgetItem([f'All The Agents  : {len(agentsForTrip)}'])
        #         for agent in agentsForTrip:
        #             rr = QtWidgets.QTreeWidgetItem([str(i) for i in agent])
        #             ch1.addChild(rr)
        #         item.addChild(ch1)
        #     # if f'{datetime.date.today()}' in  i[3] :
        #     #     ##print(self.trips.topLevelItem(ind))
        #
        #     self.trips.addTopLevelItem(item)
        #     ind += 1
        cnx.close()


class Trip_View(QtWidgets.QWidget):
    def __init__(self, role, id, desibleedit = False, p=False):
        super(Trip_View, self).__init__()
        QtWidgets.QWidget.__init__(self)
        uic.loadUi(os.path.join(os.path.dirname(__file__), "ui/tripView.ui"), self)
        self.role = role
        if self.role == 0:
            print('The Admin')
        # ##print(data)
        self.id_ = id
        self.p = p
        cnx = con()
        cur = cnx.cursor()
        cur.execute(f'''select v.matr, t.datetime, concat(d.firstName, " ", d.LastName) as driverName, t.starttime, t.startloc, t.stoptime, t.stoploc, t.tracking from
                        vans v inner join trips t on t.van = v.id inner join drivers d on t.driver = d.id where t.id = {int(self.id_)}''')
        self.setWindowTitle('View The Trip\'s Details')
        self.trip_id.setText(str(self.id_))

        dt = cur.fetchone()

        self.van_matr.setText(dt[0])
        self.trip_time.setText(dt[1])
        self.driver_name.setText(dt[2])

        self.start_time.setText(dt[3])

        self.start_time.setToolTip(f"Click to View Start location on Google maps {dt[4]}")

        self.start_loc = dt[4]
        self.stop_time.setText(dt[5])

        self.stop_time.setToolTip(f"Click to View Stop location on Google maps {dt[6]}")
        self.trck.setToolTip(f"Click to View the tracking points for this trip .")

        self.stop_loc = dt[6]
        self.trajet = ""
        if dt[7]:
            print(dt[7])
            litlot = json.loads(dt[7].replace("'", '"'))
            print("#"*100)
            print(type(litlot),litlot)
            for i in litlot.values():
                print(i)
                self.trajet += f'/{i}'



        self.rmv.clicked.connect(self.removeAgent)

        self.tableWidget.itemSelectionChanged.connect(self.rowSelected)
        self.agnts.currentIndexChanged.connect(lambda : self.add.setEnabled(True) if self.agnts.currentIndex() != 0 else self.add.setEnabled(False))

        cnx.close()

        self.add.clicked.connect(self.add_agent)
        self.readOnly = desibleedit
        if self.readOnly:
            self.agnts.setEnabled(False)
            self.tableWidget.setEnabled(False)
        self.prnt.clicked.connect(self.printTrip)




        self.refreshTable()
        self.back_icon.setPixmap(QtGui.QPixmap('src/img/back.png'))
        self.back_icon.setScaledContents(True)
        self.back_icon.installEventFilter(self)
        self.cls.setPixmap(QtGui.QPixmap('src/img/cls.png'))
        self.cls.setScaledContents(True)
        self.cls.installEventFilter(self)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.offset = None

        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.start_time.installEventFilter(self)
        self.stop_time.installEventFilter(self)
        self.trck.installEventFilter(self)

    def openMap(self, latLot):
        if latLot and len(latLot.split(",")) == 2:
            lat, lot = [i for i in str(latLot).split(',')]
            webbrowser.open(f'''http://maps.google.com/maps?z=12&t=m&q=loc:{lat}+{lot}''')

    def paintEvent(self, event):
        # get current window size
        opt = QtWidgets.QStyleOption()
        opt.initFrom(self)
        rect = opt.rect

        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        p.setBrush(opt.palette.brush(QPalette.Window))
        p.setPen(QtCore.Qt.NoPen)
        p.drawRoundedRect(rect, 40, 40)
        p.end()

    def mousePressEvent(self, event):
        if event.button() == QtCore.Qt.LeftButton:
            self.offset = event.pos()
        else:
            super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self.offset is not None and event.buttons() == QtCore.Qt.LeftButton:
            self.move(self.pos() + event.pos() - self.offset)
        else:
            super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self.offset = None
        super().mouseReleaseEvent(event)

    def eventFilter(self, s, e):
        if s is self.back_icon:
            if e.type() == QtCore.QEvent.MouseButtonPress:
                self.close()
            if e.type() == QtCore.QEvent.Enter:
                self.back_icon.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                ##print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.back_icon.setStyleSheet('border-color: 0px ;')
                ##print('mouse out')
        if s is self.cls:
            if e.type() == QtCore.QEvent.Enter:
                self.cls.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                ##print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.cls.setStyleSheet('border-color: 0px ;')
                ##print('mouse out')
            elif e.type() == QtCore.QEvent.MouseButtonPress:
                os._exit(0)

        if s is self.start_time and e.type() == QtCore.QEvent.MouseButtonPress:
                self.openMap(self.start_loc)

        if s is self.stop_time and e.type() == QtCore.QEvent.MouseButtonPress:
                self.openMap(self.stop_loc)

        if s is self.trck and e.type() == QtCore.QEvent.MouseButtonPress:
                webbrowser.open(f"https://www.google.com/maps/dir{self.trajet}")

        return super(Trip_View, self).eventFilter(s, e)

    def printTrip(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        fileName, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save Trip Planning Excel File ...", f"Trip{self.trip_id.text()}_{self.trip_time.text().replace(':', '')}_{self.driver_name.text()}_{self.van_matr.text()}.xlsx",
                                                  "All Files (*);;Text Files (*.txt)")
        if fileName:

            workbook = load_workbook(filename=os.path.join(os.getcwd(), 'src', 'template.xlsx'))
            sheet = workbook.active

            cnx = con()
            cur = cnx.cursor()
            cur.execute(
                f'''select a.firstName, a.LastName, g.name from 
                trips_history th 
                inner join agents a on th.agent = a.id 
                inner join trips t on th.trip = t.id 
                inner join grps g on a.grp = g.id  
                where t.id = {self.trip_id.text()} order by g.name asc''')
            data = cur.fetchall()
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            sheet['D1'] = f'{self.trip_id.text()} / {self.driver_name.text()}'
            sheet['D2'] = f'{self.trip_time.text()}'

            cc = [i for i in 'B C D '.split()]
            indx = 1
            for i in range(4, len(data) + 4):
                sheet[f'A{i}'] = indx
                indx += 1
                for c in range(len(cc)):
                    sheet[f"{cc[c]}{i}"] = data[i-4][c]

                    ##print(f"{cc[c]}{i}")

            for r in range(1, len(data) + 5):
                for c in range(1, 8):
                    sheet.cell(row=r, column=c).border = thin_border

            workbook.save(filename=fileName)
            os.startfile(fileName)
        else:
            notif(title='Notice ', msg='The printing Operation Has been Canceled .')

    def add_agent(self):

        cnx = con()
        cur = cnx.cursor()
        ##print(str(self.agnts.currentText()).split('-'))
        fname, lname = str(self.agnts.currentText()).split('-')
        ##print(f'First name ==> {fname}\nLast Name ==> {lname}')

        cur.execute(f'insert into trips_history(trip, agent, presence) values({int(self.trip_id.text())}, (select id from agents where firstName like "{fname}" and LastName like "{lname}"), 0) ')
        cnx.commit()
        cnx.close()
        self.refreshTable()

    def removeAgent(self):
        if items := [i.text() for i in self.tableWidget.selectedItems()]:
            ##print(items)
            reply = QtWidgets.QMessageBox.question(self, 'Alert !!', 'Are you sure you want to remove ?',
                                         QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)
            if reply == QtWidgets.QMessageBox.Yes:
                ##print('removing')
                cnx = con()
                cur = cnx.cursor()
                cur.execute(f'delete from trips_history where trip = {int(self.trip_id.text())} and agent = {int(items[0])}')
                cnx.commit()
                cnx.close()
                self.refreshTable()



    def refreshTable(self):

        [self.tableWidget.removeRow(0) for _ in range(self.tableWidget.rowCount())]
        cnx = con()
        cur = cnx.cursor()
        cur.execute(f'''select max_places from vans where matr like "{self.van_matr.text()}";''')
        # if int(self.tableWidget.rowCount()) == int(cur.fetchone()[0]) :
        #     self.agnts.setEnabled(False)
        # else:
        #     self.agnts.setEnabled(True)

        cur.execute(f'''select a.id, concat(a.firstName, " ", a.LastName) as fullName, g.name, g.shift, a.address 
                                    from trips_history th inner join agents a on th.agent = a.id 
                                    inner join grps g on a.grp = g.id where th.trip = {int(self.id_)};''')
        data = [[c for c in r] for r in cur.fetchall()]
        cur.execute(f'select max_places from vans where matr like "{self.van_matr.text()}"')
        maxP = int(cur.fetchone()[0])

        if agents := data:
            if not self.readOnly:
                if len(agents) >= maxP :
                    self.agnts.setEnabled(False)
                else:
                    self.agnts.setEnabled(True)

            self.tableWidget.setRowCount(len(agents))
            self.tableWidget.setColumnCount(len(agents[0]))
            self.tableWidget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
            self.tableWidget.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            self.tableWidget.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
            self.tableWidget.horizontalHeader().setSectionResizeMode(3, QHeaderView.Fixed)
            self.tableWidget.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)



            head = [i for i in 'Agent ID-Full Name-Group-Shift-Address'.split('-')]
            self.tableWidget.setHorizontalHeaderLabels(head)
            ##print(agents)
            for r in range(len(agents)):
                for c in range(len(agents[r])):
                    self.tableWidget.setItem(r, c, QtWidgets.QTableWidgetItem(str(agents[r][c])))

        cur.execute(f'select agent from trips_history where trip = {self.trip_id.text()}')
        # ##print(f'current agents == > {list(set([i[0] for i in cur.fetchall()]))}')
        agents = []
        dtt = cur.fetchall()
        ##print(dtt)
        query = ''
        if dtt:
            existsAgents = tuple(set([i[0] for i in dtt])) if len(dtt) > 1 else (dtt[0]) if dtt else "()"
            if len(existsAgents) > 1:
                query = f'select concat(firstName, "-", LastName) as fullName from agents where id not in {existsAgents};'
            else:
                query = f'select concat(firstName, "-", LastName) as fullName from agents where id != {existsAgents[0]};'

        else:
            query = f'select concat(firstName, "-", LastName) as fullName from agents;'

        cur.execute(query= query)
        self.agnts.clear()
        self.agnts.addItems(['Choose Agent ...'])
        self.agnts.addItems([i[0] for i in cur.fetchall()])
        self.agnts.setCurrentIndex(0)

        cnx.close()

        ##print('reched finiched')


    def rowSelected(self):
        items = [i.text() for i in self.tableWidget.selectedItems()]
        ##print(items)
        if items:
            self.rmv.setEnabled(True)
        else:
            self.rmv.setEnabled(False)

    def closeEvent(self, event):
        if self.p:
            self.bb = Trips(self.role)
        else:
            self.bb = Main(self.role)
        self.bb.show()
        self.close()


class AddTrip(QtWidgets.QWidget):
    def __init__(self, role):
        super(AddTrip, self).__init__()
        QtWidgets.QWidget.__init__(self)
        uic.loadUi(os.path.join(os.path.dirname(__file__), "ui/addTrip.ui"), self)
        self.save_btn.clicked.connect(self.save_)
        self.role = role
        if self.role == 0:
            print('The Admin')



        self.back_icon.setPixmap(QtGui.QPixmap('src/img/back.png'))
        self.back_icon.setScaledContents(True)
        self.back_icon.installEventFilter(self)
        # cur_date_time = strftime("%Y-%MM-%dd %hh:%mm:%ss", gmtime())
        # now = QtCore.QDateTime.fromString(cur_date_time, 'yyyy-MM-dd hh:mm:ss')f
        # ##print(now)
        now = datetime.datetime.now()
        self.dateTimeEdit.setDateTime(now)
        ##print(now == self.dateTimeEdit.dateTime())
        self.dateTimeEdit.dateTimeChanged.connect(lambda : self.save_btn.setEnabled(True) if datetime.datetime.now() <= self.dateTimeEdit.dateTime() else self.save_btn.setEnabled(False))
        cnx = con()
        cur = cnx.cursor()
        cur.execute('select concat(firstName, " ", LastName) as fullName from drivers;')
        self.drivers_.addItems([i[0] for i in cur.fetchall()])
        cur.execute('select matr from vans;')
        self.van_.addItems([i[0] for i in cur.fetchall()])
        self.back = True
        cnx.close()

    def eventFilter(self, s, e):
        if s is self.back_icon:
            if e.type() == QtCore.QEvent.MouseButtonPress:
                self.close()
            if e.type() == QtCore.QEvent.Enter:
                self.back_icon.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                ##print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.back_icon.setStyleSheet('border-color: 0px ;')
                #print('mouse out')
        return super(AddTrip, self).eventFilter(s, e)

    def closeEvent(self, event):
        if self.back:
            self.TripsPage = Trips(self.role)
            self.TripsPage.show()
            self.close()


    def save_(self):
        dt = self.dateTimeEdit.dateTime().toString(("yyyy-MM-dd HH"))
        print(dt)
        if datetime.datetime.now() <= self.dateTimeEdit.dateTime():
            # cnx = con()
            # cur = cnx.cursor()
            #
            # cur.execute(f'select id from trips where datetime like "{dt}%"')
            # trip = cur.fetchone()
            # if not trip:
            #
            #     cur.execute(f'''insert into trips(van, driver, datetime, ttype) values(
            #     (select id from vans where matr like "{self.van_.currentText()}"),
            #     (select id from drivers where firstName like "{str(self.drivers_.currentText()).split(" ")[0]}" and LastName like "{str(self.drivers_.currentText()).split(" ")[1]}"),
            #      "{dt}:00:00",
            #       "{"IN" if self.IN.isChecked() else "OUT"}"
            #       )''')
            #     cnx.commit()
            #
            # #print(f'DT = {dt}')
            # cur.execute(f'select id from trips where datetime like "{dt}%"')
            # id = [cur.fetchone()[0]]
            # # #print(id)
            # self.tripView = Trip_View(id=id, role=self.role)
            # self.tripView.show()
            # self.back = False
            # self.close()
            #
            #
            # cnx.close()
            create_trip(date=f'{dt}:00:00', van=self.van_.currentText(), type="IN" if self.IN.isChecked() else "OUT", driver=self.drivers_.currentText())
        else:
            notif(title="ERORR", msg=f"Can't create a trip with this date : {dt}")


class Trips(QtWidgets.QWidget):
    def __init__(self, role):
        super(Trips, self).__init__()
        QtWidgets.QWidget.__init__(self)
        uic.loadUi(os.path.join(os.path.dirname(__file__), "ui/trips.ui"), self)
        self.role = role
        self.getTrips()
        # self.create_btn.clicked.connect(self.createTrips)
        self.tableWidget.doubleClicked.connect(self.openTripView)
        self.back = True
        self.setWindowTitle('All Trips')


        self.back_icon.setPixmap(QtGui.QPixmap('src/img/back.png'))
        self.back_icon.setScaledContents(True)
        self.back_icon.installEventFilter(self)
        self.create_btn.clicked.connect(self.createManualTrip)
        self.cls.setPixmap(QtGui.QPixmap('src/img/cls.png'))
        self.cls.setScaledContents(True)
        self.cls.installEventFilter(self)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.offset = None
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)


    def paintEvent(self, event):
        # get current window size
        opt = QtWidgets.QStyleOption()
        opt.initFrom(self)
        rect = opt.rect

        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        p.setBrush(opt.palette.brush(QPalette.Window))
        p.setPen(QtCore.Qt.NoPen)
        p.drawRoundedRect(rect, 40, 40)
        p.end()

    def mousePressEvent(self, event):
            if event.button() == QtCore.Qt.LeftButton:
                self.offset = event.pos()
            else:
                super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
            if self.offset is not None and event.buttons() == QtCore.Qt.LeftButton:
                self.move(self.pos() + event.pos() - self.offset)
            else:
                super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
            self.offset = None
            super().mouseReleaseEvent(event)


    def createManualTrip(self):
        self.back = False
        self.addTripPage = AddTrip(role=self.role)
        self.addTripPage.show()
        self.close()

    def eventFilter(self, s, e):
        if s is self.back_icon:
            if e.type() == QtCore.QEvent.MouseButtonPress:
                self.close()
            if e.type() == QtCore.QEvent.Enter:
                self.back_icon.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                #print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.back_icon.setStyleSheet('border-color: 0px ;')
                #print('mouse out')

        if s is self.cls:
            if e.type() == QtCore.QEvent.Enter:
                self.cls.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                #print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.cls.setStyleSheet('border-color: 0px ;')
                #print('mouse out')
            elif e.type() == QtCore.QEvent.MouseButtonPress:
                os._exit(0)

        return super(Trips, self).eventFilter(s, e)

    def closeEvent(self, event):
        if self.back:
            self.main = Main(self.role)
            self.main.show()
            self.close()

    # def createTrips(self):
    #     today = datetime.date.today()
    #     day = datetime.datetime.today().strftime('%A')
    #     #print(day)
    #     if day.lower() == "Friday".lower():
    #         tomorrow = datetime.date.today() + datetime.timedelta(days=3)
    #     else:
    #         tomorrow = datetime.date.today() + datetime.timedelta(days=1)
    #
    #     # tomorrow = today
    #     #print(tomorrow)
    #     cnx = con()
    #     cur = cnx.cursor()
    #     cur.execute('select shift from grps')
    #     data = []
    #     for r in cur.fetchall():
    #         for c in r:
    #             for cc in str(c).split('-'):
    #                 data.append(int(cc))
    #     dt = [i for i in set(sorted(data))]
    #     data = []
    #     for i in dt:
    #         if i < 10:
    #             data.append(f'0{i}')
    #         else:
    #             data.append(str(i))
    #     # #print(data)
    #     trips = {}
    #     total = 0
    #     ids = []
    #     for h in data:
    #         query = ''
    #         if int(h) < 15:
    #             query = f'select a.id from agents a inner join grps g on a.grp = g.id where g.shift like "{h}-%"'
    #         else:
    #             query = f'select a.id from agents a inner join grps g on a.grp = g.id where g.shift like "%-{h}"'
    #
    #         cur.execute(query)
    #         result = cur.fetchall()
    #         # #print(result)
    #
    #         count = len(result)
    #
    #         trips[h] = [i[0] for i in result]
    #         total += count
    #
    #     #print(trips)
    #     #print(total)
    #     maxForVan = 18
    #     for time, agents in trips.items():
    #         vansNeeded = 0
    #         if agents:
    #             if len(agents) > maxForVan:
    #                 vansNeeded = 2
    #             else:
    #                 vansNeeded = 1
    #
    #             if vansNeeded == 1:
    #                 cur.execute(
    #                     f''' select count(id) from trips where datetime like "{tomorrow} {time}:00:00" and van = 1''')
    #                 if not cur.fetchone()[0]:
    #                     cur.execute(
    #                         f'''insert into trips(van, driver, datetime) values (1, 1, "{tomorrow} {time}:00:00");''')
    #                     cnx.commit()
    #                     for agent in agents:
    #                         cur.execute(f'''select id from trips where datetime like "{tomorrow} {time}:00:00"''')
    #                         cur.execute(
    #                             f'''insert into trips_history (trip, agent, presence) values ({int(cur.fetchone()[0])}, {agent}, 0)''')
    #                         cnx.commit()
    #             else:
    #                 grp1 = [i for i in agents[:len(agents) // 2]]
    #                 grp2 = [i for i in agents[len(agents) // 2:]]
    #                 cur.execute(
    #                     f'''select count(id) from trips where datetime like "{tomorrow} {time}:00:00" and driver = 1''')
    #                 if not cur.fetchone()[0]:
    #                     cur.execute(
    #                         f'''insert into trips(van, driver, datetime) values (1, 1, "{tomorrow} {time}:00:00");''')
    #
    #                 cur.execute(f'''select id from trips where datetime like "{tomorrow} {time}:00:00"''')
    #
    #                 trip1, trip2 = [int(i[0]) for i in cur.fetchall()]
    #                 for agent in grp1:
    #                     cur.execute(f'''select count(id) from trips_history where trip = {trip1} and agent = {agent}''')
    #                     if not cur.fetchone()[0]:
    #                         cur.execute(
    #                             f'''insert into trips_history (trip, agent, presence) values ({trip1}, {agent}, 0)''')
    #                         cnx.commit()
    #
    #                 cur.execute(
    #                     f'''select count(id) from trips where datetime like "{tomorrow} {time}:00:00" and driver = 1''')
    #                 if not cur.fetchone()[0]:
    #                     cur.execute(
    #                         f'''insert into trips(van, driver, datetime) values (2, 2, "{tomorrow} {time}:00:00");''')
    #                 cnx.commit()
    #                 for agent in grp2:
    #                     cur.execute(f'''select count(id) from trips_history where trip = {trip2} and agent = {agent}''')
    #                     if not cur.fetchone()[0]:
    #                         cur.execute(
    #                             f'''insert into trips_history (trip, agent, presence) values ({trip2}, {agent}, 0)''')
    #                         cnx.commit()
    #     cnx.close()
    #     self.getTrips()
    #


    def getTrips(self):
            self.tableWidget.setRowCount(0)
            cnx = con()
            cur = cnx.cursor()
            cur.execute('''select t.id, v.matr, d.firstName, d.LastName, t.datetime, (select count(id) from trips_history where trip = t.id) as Agent_numbers
                    from trips t inner join vans v on t.van = v.id inner join drivers d on t.driver = d.id order by t.datetime desc;''')
            dt = cur.fetchall()

            data = []

            for r in dt:
                data.append([r[0], r[1], f'{r[2]} {r[3]}', r[4], r[5]])
            head = ['Trip-ID ', 'van Matrecule', 'Driver Name', 'Time', 'Agents Numbers']

            #print(data)
            self.tableWidget.setColumnCount(len(data[0]))
            self.tableWidget.setRowCount(len(data))
            # self.tableWidget.horizontalHeader().setSectionResizeMode(head.index(head[-1]), QHeaderView.Stretch)
            # self.tableWidget.resizeColumnsToContents()
            for i in range(self.tableWidget.columnCount()):
                self.tableWidget.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)

            self.tableWidget.setHorizontalHeaderLabels(head)
            for r in range(len(data)):
                for c in range(len(data[0])):
                    self.tableWidget.setItem(r, c, QtWidgets.QTableWidgetItem(str(data[r][c])))

            cnx.close()

    def openTripView(self):
        self.back = False
        #print(self.tableWidget.currentIndex().siblingAtColumn(0).data())

        data = [self.tableWidget.currentIndex().siblingAtColumn(i).data() for i in range(self.tableWidget.columnCount())]
        #print(f'data ==> {data}')
        cnx = con()
        cur = cnx.cursor()
        cur.execute(f'select datetime from trips where id = {int(data[0])} ;')
        dd = cur.fetchone()[0]

        trDate = datetime.datetime.strptime(f"{str(dd).split(' ')[0]} {str(dd).split(' ')[1].split(':')[0]}",
                                            '%Y-%m-%d %H')
        now = datetime.datetime.strptime(datetime.datetime.today().strftime('%Y-%m-%d %H'), '%Y-%m-%d %H')

        # if int(strftime("%H", gmtime())) >= int(str(dd).split(' ')[1].split(':')[0]) and int(str(dd).split(' ')[0].split('-')[2]) < int(strftime("%d", gmtime())) and int(str(dd).split(' ')[0].split('-')[1]) <= int(strftime("%m", gmtime())) :
        if now > trDate:
            self.TV = Trip_View(role=self.role, id=data[0], desibleedit=True, p=True)
            notif(self, title="Error", msg='You can\'t Modify this Trip .')
        else:
            self.TV = Trip_View(role=self.role, id=data[0], p=True)
        self.TV.show()
        self.close()


class Agents(QtWidgets.QWidget):
    def __init__(self, role):
        super(Agents, self).__init__()
        QtWidgets.QWidget.__init__(self)
        uic.loadUi(os.path.join(os.path.dirname(__file__), "ui/Agents.ui"), self)
        self.add_btn.clicked.connect(self.add)
        self.setWindowTitle('All Agents')
        self.role = role
        if self.role == 0:
            print("The Admin")

        self.refreshTable()

        self.back = True

        self.edit_btn.clicked.connect(self.edit)

        self.groups.currentIndexChanged.connect(self.refreshTable)

        self.tableWidget.itemSelectionChanged.connect(self.tableSelectRowChanged)
        self.search.textChanged.connect(self.searching)


        self.back_icon.setPixmap(QtGui.QPixmap('src/img/back.png'))
        self.back_icon.setScaledContents(True)
        self.back_icon.installEventFilter(self)
        self.cls.setPixmap(QtGui.QPixmap('src/img/cls.png'))
        self.cls.setScaledContents(True)
        self.cls.installEventFilter(self)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.offset = None

        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

    def paintEvent(self, event):
        # get current window size
        opt = QtWidgets.QStyleOption()
        opt.initFrom(self)
        rect = opt.rect

        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        p.setBrush(opt.palette.brush(QPalette.Window))
        p.setPen(QtCore.Qt.NoPen)
        p.drawRoundedRect(rect, 40, 40)
        p.end()

    def mousePressEvent(self, event):
            if event.button() == QtCore.Qt.LeftButton:
                self.offset = event.pos()
            else:
                super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
            if self.offset is not None and event.buttons() == QtCore.Qt.LeftButton:
                self.move(self.pos() + event.pos() - self.offset)
            else:
                super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
            self.offset = None
            super().mouseReleaseEvent(event)



    def eventFilter(self, s, e):
        if s is self.back_icon:
            if e.type() == QtCore.QEvent.MouseButtonPress:
                self.close()
            if e.type() == QtCore.QEvent.Enter:
                self.back_icon.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                #print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.back_icon.setStyleSheet('border-color: 0px ;')
                #print('mouse out')
        if s is self.cls:
            if e.type() == QtCore.QEvent.Enter:
                self.cls.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                #print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.cls.setStyleSheet('border-color: 0px ;')
                #print('mouse out')
            elif e.type() == QtCore.QEvent.MouseButtonPress:
                os._exit(0)

        return super(Agents, self).eventFilter(s, e)

    def searching(self):

        self.refreshTable(key= self.search.text() if self.search.text() else None)

    def refreshTable(self, key = None):

        self.tableWidget.setRowCount(0)

        cnx = con()
        cur = cnx.cursor()

        if key:= self.search.text():
            query = f'''select a.matrr, a.firstName, a.LastName, a.CIN, a.address, g.name, g.shift, a.zone  from 
            agents a inner join grps g on a.grp = g.id where
            {f"(g.name like '{self.groups.currentText()}') and "if self.groups.currentIndex() != 0 else " "} ( a.firstName like "%{key}%" or a.LastName like "%{key}%" or a.CIN like "%{key}%" or a.address like "%{key}%" or g.name like "%{key}%");'''
        else:
            query = f'''select a.matrr, a.firstName, a.LastName, a.CIN, a.address, g.name, g.shift, a.zone  from 
                        agents a inner join grps g on a.grp = g.id {f" where g.name like '{self.groups.currentText()}'" if self.groups.currentIndex() != 0 else ""}'''

        #print('#'*30)
        #print(key)

        #print(query)
        # if self.groups.currentIndex() == 0:
        #
        #
        #     # self.search.setEnabled(True)
        #     self.search.setStyleSheet('border : 1px solid blue')
        #     if key:
        #         cur.execute(
        #             f'select {clms} from agents a inner join grps g on a.grp = g.id where a.firstName like "%{key}%" or a.LastName like "%{key}%" or a.CIN like "%{key}%" or a.address like "%{key}%" or g.name like "%{key}%";')
        #     else:
        #         cur.execute(
        #         f'select a.id, a.firstName, a.LastName, a.CIN, a.address, g.name, g.shift from agents a inner join grps g on a.grp = ;')
        # else:
        #     # self.search.setEnabled(False)
        #     # self.search.setStyleSheet('border : 2px solid red')
        #     cur.execute(f'select id from grps where name = "{self.groups.currentText()}"')
        #     cur.execute(
        #         f'select a.id, a.firstName, a.LastName, a.CIN, a.address, g.name, g.shift from agents a inner join grps g on a.grp = g.id  where a.grp = {cur.fetchone()[0]} and (a.firstName like "%{key}%" or a.LastName like "%{key}%" or a.CIN like "%{key}%" or a.address like "%{key}%");')


        agents = []
        cur.execute(query)
        data = cur.fetchall()

        if data:


            for r in data:
                agents.append([r[0], f'{r[1]} {r[2]}', r[3], r[4], r[5], r[6], r[7]])
            head = ['Agent ID ', 'Full Name', 'CIN', 'Adress', 'Group', 'Shift', 'Zone']
            #print(agents)
            self.tableWidget.setColumnCount(len(agents[0]))
            self.tableWidget.setRowCount(len(agents))
            # self.tableWidget.horizontalHeader().setSectionResizeMode(head.index(head[-1]), QHeaderView.Stretch)
            # self.tableWidget.resizeColumnsToContents()

            self.tableWidget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
            self.tableWidget.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            self.tableWidget.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
            self.tableWidget.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
            self.tableWidget.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
            self.tableWidget.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
            self.tableWidget.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)

            self.tableWidget.setHorizontalHeaderLabels(head)
            for r in range(len(agents)):
                for c in range(len(agents[0])):
                    self.tableWidget.setItem(r, c, QtWidgets.QTableWidgetItem(str(agents[r][c])))

            cur.execute('select name from grps')
            #print(f'Groups Items : {[self.groups.itemText(i) for i in range(self.groups.count())]}')
            grps_names = []
            for o in cur.fetchall():
                if o[0] not in [self.groups.itemText(i) for i in range(self.groups.count())]:
                    grps_names.append(o[0])



            #print(grps_names)
            # self.groups.clear()
            self.groups.addItems([str(i) for i in grps_names])
        else:
            print('no data found ')
        cnx.close()


    def tableSelectRowChanged(self):
        items = [i.text() for i in self.tableWidget.selectedItems()]
        #print(items)
        if items:
            self.edit_btn.setEnabled(True)
        else:
            self.edit_btn.setEnabled(False)


    def edit(self):

        self.back = False
        dt = [i.text() for i in self.tableWidget.selectedItems()]
        cnx = con()
        cur = cnx.cursor()
        cur.execute(f'select street from agents where matrr = "{dt[0]}"')
        dt.append(cur.fetchone()[0])
        print(f'Data ==> {dt}')
        self.eA = AddAgent(self.role, do='edit', data=dt)
        self.eA.show()
        self.close()


    def closeEvent(self, event):
        if self.back:
            self.main = Main(self.role)
            self.main.show()
            self.close()

    def add(self):
        self.back = False
        self.addAgent = AddAgent(self.role)
        self.close()
        self.addAgent.show()


class Vans(QtWidgets.QWidget):
    def __init__(self, role):
        super(Vans, self).__init__()
        QtWidgets.QWidget.__init__(self)
        uic.loadUi(os.path.join(os.path.dirname(__file__), "ui/vans.ui"), self)
        self.add_btn.clicked.connect(self.add)
        self.setWindowTitle('All vehicles')
        self.role = role
        if self.role == 0:
            print("The Admin")

        self.refreshTable()

        self.back = True

        self.edit_btn.clicked.connect(self.edit)


        self.tableWidget.itemSelectionChanged.connect(self.tableSelectRowChanged)


        self.back_icon.setPixmap(QtGui.QPixmap('src/img/back.png'))
        self.back_icon.setScaledContents(True)
        self.back_icon.installEventFilter(self)
        self.cls.setPixmap(QtGui.QPixmap('src/img/cls.png'))
        self.cls.setScaledContents(True)
        self.cls.installEventFilter(self)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.offset = None

        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

    def paintEvent(self, event):
        # get current window size
        opt = QtWidgets.QStyleOption()
        opt.initFrom(self)
        rect = opt.rect

        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        p.setBrush(opt.palette.brush(QPalette.Window))
        p.setPen(QtCore.Qt.NoPen)
        p.drawRoundedRect(rect, 40, 40)
        p.end()


    def mousePressEvent(self, event):
            if event.button() == QtCore.Qt.LeftButton:
                self.offset = event.pos()
            else:
                super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
            if self.offset is not None and event.buttons() == QtCore.Qt.LeftButton:
                self.move(self.pos() + event.pos() - self.offset)
            else:
                super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
            self.offset = None
            super().mouseReleaseEvent(event)



    def eventFilter(self, s, e):
        if s is self.back_icon:
            if e.type() == QtCore.QEvent.MouseButtonPress:
                self.close()
            if e.type() == QtCore.QEvent.Enter:
                self.back_icon.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                #print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.back_icon.setStyleSheet('border-color: 0px ;')
                #print('mouse out')

        if s is self.cls:
            if e.type() == QtCore.QEvent.Enter:
                self.cls.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                #print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.cls.setStyleSheet('border-color: 0px ;')
                #print('mouse out')
            elif e.type() == QtCore.QEvent.MouseButtonPress:
                os._exit(0)

        return super(Vans, self).eventFilter(s, e)

    def searching(self):
        self.refreshTable(key= self.search.text())

    def refreshTable(self, key = None):

        self.tableWidget.setRowCount(0)

        cnx = con()
        cur = cnx.cursor()
        columns = 'v.id, v.matr, v.max_places, concat(d.firstName, " ", LastName) as fullName '

        cur.execute(
            f'select {columns} from vans v inner join drivers d on v.driver = d.id ;')


        data = cur.fetchall()

        if data:
            vans = [[c for c in r] for r in data]
            head = ['Vehicle ID ', 'Matricule', 'Max Places', 'Driver\'s name']
            #print(vans)
            self.tableWidget.setColumnCount(len(vans[0]))
            self.tableWidget.setRowCount(len(vans))
            # self.tableWidget.horizontalHeader().setSectionResizeMode(head.index(head[-1]), QHeaderView.Stretch)
            # self.tableWidget.resizeColumnsToContents()
            for i in range(self.tableWidget.columnCount()):
                self.tableWidget.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)

            self.tableWidget.setHorizontalHeaderLabels(head)
            for r in range(len(vans)):
                for c in range(len(vans[0])):
                    self.tableWidget.setItem(r, c, QtWidgets.QTableWidgetItem(str(vans[r][c])))


        else:
            print('no data found ')
        cnx.close()


    def tableSelectRowChanged(self):
        items = [i.text() for i in self.tableWidget.selectedItems()]
        #print(items)
        if items:
            self.edit_btn.setEnabled(True)
        else:
            self.edit_btn.setEnabled(False)



    def closeEvent(self, event):
        if self.back:
            self.main = Main(self.role)
            self.main.show()
            self.close()

    def add(self):
        self.back = False
        self.addVans = AddVan(self.role)
        self.close()
        self.addVans.show()

    def edit(self):

        self.back = False
        self.eV = AddVan(self.role, do='edit', data=[i.text() for i in self.tableWidget.selectedItems()])
        self.eV.show()
        self.close()


class AddVan(QtWidgets.QWidget):
    def __init__(self, role, do = 'save', data = None):
        super(AddVan, self).__init__()
        QtWidgets.QWidget.__init__(self)
        uic.loadUi(os.path.join(os.path.dirname(__file__), "ui/addVan.ui"), self)
        self.save_btn.clicked.connect(self.save)
        self.role = role
        if self.role == 0:
            print('The Admin')
        self.action = do
        cnx = con()
        cur = cnx.cursor()
        cur.execute('select concat(firstName, " ", LastName) as fullName from drivers;')
        drivers_names = [i[0] for i in cur.fetchall()]
        #print(drivers_names)
        self.drivers.addItems(drivers_names)
        self.data = data
        if self.data:
            self.matr.setText(str(self.data[1]))
            self.max_places.setText(str(self.data[2]))
            self.drivers.setCurrentIndex(drivers_names.index(self.data[3]))
        self.setWindowTitle('Add new Vehicle')


        cnx.close()
        self.back_icon.setPixmap(QtGui.QPixmap('src/img/back.png'))
        self.back_icon.setScaledContents(True)
        self.back_icon.installEventFilter(self)

    def eventFilter(self, s, e):
        if s is self.back_icon:
            if e.type() == QtCore.QEvent.MouseButtonPress:
                self.close()
            if e.type() == QtCore.QEvent.Enter:
                self.back_icon.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                #print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.back_icon.setStyleSheet('border-color: 0px ;')
                #print('mouse out')
        return super(AddVan, self).eventFilter(s, e)

    def closeEvent(self, event):
        self.VansPage = Vans(self.role)
        self.VansPage.show()
        self.close()



    def save(self):
        if len(self.matr.text()) > 3 and len(self.max_places.text()) > 0:
            cnx = con()
            cur = cnx.cursor()
            cur.execute(f'select id from drivers where firstName like "{str(self.drivers.currentText()).split(" ")[0]}" and LastName like "{str(self.drivers.currentText()).split(" ")[-1]}"')
            drvr = int(cur.fetchone()[0])

            if self.action == 'save':
                cur.execute(
                    f'select count(id) from vans where matr like "{self.matr.text()}";')
                if cur.fetchone()[0]:
                    #print('this Van alredy exists ')
                    self.matr.setStyleSheet('border : 2px solid red')
                    notif(self, title='Transport Planning ', msg='This Vehicle already exists in the Database ')
                else:
                    cur.execute(f'insert into vans (matr, max_places, driver) value ("{self.matr.text()}", {int(self.max_places.text())}, {drvr});')
                    cnx.commit()
            elif self.action == 'edit':
                    cur.execute(f'''
                            update vans set matr = "{self.matr.text()}",
                                            max_places = {int(self.max_places.text())},
                                            
                                            driver = {drvr}
                                            where id = {int(self.data[0])}
                        ''')
                    cnx.commit()
                    # self.agents = Agents(self.role)
                    # self.agents.show()
                    # self.close()
            self.matr.clear()
            self.max_places.clear()
            self.drivers.setCurrentIndex(0)

            cnx.close()

        else:
            notif(self, title="Error ", msg='you have to write all the infos')


class Compaigns(QtWidgets.QWidget):
    def __init__(self, role):
        super(Compaigns, self).__init__()
        QtWidgets.QWidget.__init__(self)
        uic.loadUi(os.path.join(os.path.dirname(__file__), "ui/cmps.ui"), self)
        self.add_btn.clicked.connect(self.add)

        self.role = role
        if self.role == 0:
            print("The Admin")

        self.refreshTable()

        self.back = True

        self.edit_btn.clicked.connect(self.edit)
        self.setWindowTitle('All Groups')
        self.tableWidget.itemSelectionChanged.connect(self.tableSelectRowChanged)

        self.back_icon.setPixmap(QtGui.QPixmap('src/img/back.png'))
        self.back_icon.setScaledContents(True)
        self.back_icon.installEventFilter(self)
        self.cls.setPixmap(QtGui.QPixmap('src/img/cls.png'))
        self.cls.setScaledContents(True)
        self.cls.installEventFilter(self)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.offset = None
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

    def paintEvent(self, event):
        # get current window size
        opt = QtWidgets.QStyleOption()
        opt.initFrom(self)
        rect = opt.rect

        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)
        p.setBrush(opt.palette.brush(QPalette.Window))
        p.setPen(QtCore.Qt.NoPen)
        p.drawRoundedRect(rect, 40, 40)
        p.end()

    def mousePressEvent(self, event):
            if event.button() == QtCore.Qt.LeftButton:
                self.offset = event.pos()
            else:
                super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
            if self.offset is not None and event.buttons() == QtCore.Qt.LeftButton:
                self.move(self.pos() + event.pos() - self.offset)
            else:
                super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
            self.offset = None
            super().mouseReleaseEvent(event)



    def eventFilter(self, s, e):
        if s is self.back_icon:
            if e.type() == QtCore.QEvent.MouseButtonPress:
                self.close()
            if e.type() == QtCore.QEvent.Enter:
                self.back_icon.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                #print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.back_icon.setStyleSheet('border-color: 0px ;')
                #print('mouse out')

        if s is self.cls:
            if e.type() == QtCore.QEvent.Enter:
                self.cls.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                #print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.cls.setStyleSheet('border-color: 0px ;')
                #print('mouse out')
            elif e.type() == QtCore.QEvent.MouseButtonPress:
                os._exit(0)


        return super(Compaigns, self).eventFilter(s, e)

    def refreshTable(self, key=None):

        self.tableWidget.setRowCount(0)

        cnx = con()
        cur = cnx.cursor()

        cur.execute(
            f'select id, name, shift  from grps ;')



        if dt := [[c for c in r] for r in cur.fetchall()]:
            data = []
            for i, r in enumerate(dt):
                row = r
                cur.execute(f'select count(id) from agents where grp = {row[0]}')
                row.append(cur.fetchone()[0])
                data.append(row)

            head = ['Compaign ID ', 'Compaign Name', 'Shift', 'Agents Number']
            #print(data)
            self.tableWidget.setColumnCount(len(data[0]))
            self.tableWidget.setRowCount(len(data))
            # self.tableWidget.horizontalHeader().setSectionResizeMode(head.index(head[-1]), QHeaderView.Stretch)
            # self.tableWidget.resizeColumnsToContents()
            for i in range(self.tableWidget.columnCount()):
                self.tableWidget.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)

            self.tableWidget.setHorizontalHeaderLabels(head)
            for r in range(len(data)):
                for c in range(len(data[0])):
                    self.tableWidget.setItem(r, c, QtWidgets.QTableWidgetItem(str(data[r][c])))


        else:
            print('no data found ')
        cnx.close()

    def tableSelectRowChanged(self):
        items = [i.text() for i in self.tableWidget.selectedItems()]
        #print(items)
        if items:
            self.edit_btn.setEnabled(True)
        else:
            self.edit_btn.setEnabled(False)

    def closeEvent(self, event):
        if self.back:
            self.main = Main(self.role)
            self.main.show()
            self.close()

    def add(self):
        self.back = False
        self.addComp = AddComp(self.role)
        self.close()
        self.addComp.show()

    def edit(self):

        self.back = False
        self.eV = AddComp(self.role, do='edit', data=[i.text() for i in self.tableWidget.selectedItems()])
        self.eV.show()
        self.close()


class AddComp(QtWidgets.QWidget):
    def __init__(self, role, do='save', data=None):
        super(AddComp, self).__init__()
        QtWidgets.QWidget.__init__(self)
        uic.loadUi(os.path.join(os.path.dirname(__file__), "ui/addComp.ui"), self)
        self.save_btn.clicked.connect(self.save)
        self.role = role
        if self.role == 0:
            print('The Admin')
        self.action = do
        self.shiftfrom.addItems([f'0{i}:00' if i < 10 else f'{i}:00' for i in range(8, 24)])
        self.shiftto.addItems([f'0{i}:00' if i < 10 else f'{i}:00' for i in range(8, 24)])
        self.data = data
        #['Compaign ID ', 'Compaign Name', 'Shift', 'Agents Number']
        if self.data:
            self.cname.setText(str(self.data[1]))
            self.shiftfrom.setCurrentIndex([self.shiftfrom.itemText(i) for i in range(self.shiftfrom.count())].index(f'{str(self.data[2]).split("-")[0]}:00'))
            self.shiftto.setCurrentIndex([self.shiftto.itemText(i) for i in range(self.shiftto.count())].index(f'{str(self.data[2]).split("-")[1]}:00'))

        self.setWindowTitle('Add new Group')
        self.back_icon.setPixmap(QtGui.QPixmap('src/img/back.png'))
        self.back_icon.setScaledContents(True)
        self.back_icon.installEventFilter(self)



    def eventFilter(self, s, e):
        if s is self.back_icon:
            if e.type() == QtCore.QEvent.MouseButtonPress:
                self.close()
            if e.type() == QtCore.QEvent.Enter:
                self.back_icon.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                #print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.back_icon.setStyleSheet('border-color: 0px ;')
                #print('mouse out')
        return super(AddComp, self).eventFilter(s, e)


    def closeEvent(self, event):
        self.CompsPage = Compaigns(self.role)
        self.CompsPage.show()
        self.close()

    def save(self):
        print(f''' name > 2 ({len(self.cname.text())}) : {len(self.cname.text()) > 2} \n
                {int(self.shiftfrom.currentText().split(':')[0])} < {int(self.shiftto.currentText().split(':')[0])} :  {int(self.shiftfrom.currentText().split(':')[0]) < int(self.shiftto.currentText().split(':')[0])} \n
                {int(self.shiftto.currentText().split(':')[0])} - {int(self.shiftfrom.currentText().split(':')[0])} == 9 ({int(self.shiftto.currentText().split(':')[0]) - int(self.shiftfrom.currentText().split(':')[0])}) :  {int(self.shiftfrom.currentText().split(':')[0]) - int(self.shiftto.currentText().split(':')[0]) == 9}''')
        if len(self.cname.text()) > 2 \
                and int(self.shiftfrom.currentText().split(':')[0]) < int(self.shiftto.currentText().split(':')[0]) \
                and int(self.shiftto.currentText().split(':')[0]) - int(self.shiftfrom.currentText().split(':')[0]) == 9:
            cnx = con()
            cur = cnx.cursor()


            if self.action == 'save':
                cur.execute(
                    f'select count(id) from grps where name like "{self.cname.text()}";')
                if cur.fetchone()[0]:
                    #print('this Compaogns alredy exists ')
                    self.matr.setStyleSheet('border : 2px solid red')
                    notif(self, title='Transport Planning ', msg='This Compaign already exists in the Database ')
                else:
                    cur.execute(
                        f'insert into grps (name, shift) value ("{self.cname.text()}", "{str(self.shiftfrom.currentText()).split(":")[0]}-{str(self.shiftto.currentText()).split(":")[0]}");')
                    cnx.commit()
            elif self.action == 'edit':
                cur.execute(f'''
                            update grps set name = "{self.cname.text()}",
                                            shift = "{str(self.shiftfrom.currentText()).split(":")[0]}-{str(self.shiftto.currentText()).split(":")[0]}"
                                            where id = {int(self.data[0])}
                        ''')
                cnx.commit()
                # self.agents = Agents(self.role)
                # self.agents.show()
                # self.close()
            self.cname.clear()
            self.shiftfrom.setCurrentIndex(0)
            self.shiftto.setCurrentIndex(0)

            cnx.close()

        else:
            notif(self, title="Error ", msg='you have to write all the infos')


class AddAgent(QtWidgets.QWidget):
    def __init__(self, role, do = 'save', data = None):
        super(AddAgent, self).__init__()
        QtWidgets.QWidget.__init__(self)
        uic.loadUi(os.path.join(os.path.dirname(__file__), "ui/addAgent.ui"), self)
        self.save_btn.clicked.connect(self.save)
        self.role = role
        if self.role == 0:
            print('The Admin')
        self.action = do
        cnx = con()
        cur = cnx.cursor()
        cur.execute('select name from grps')
        grps_names = [i[0] for i in cur.fetchall()]



        #print(grps_names)
        self.grps.addItems(grps_names)
        self.data = data
        self.agent_pic_path = ""
        if self.action == "edit":
            print(f'data[0] ==> {data[0]}')
            cur.execute(f"select pic from agents where matrr like '{data[0]}'")
            self.agent_pic_path = cur.fetchone()[0]
            self.matrr.setReadOnly(True)
        else:
            self.agent_pic_path = 'https://firebasestorage.googleapis.com/v0/b/saccomxd-stm-yassine-baghdadi.appspot.com/o/profile.png?alt=media&token=30aaedee-2f7d-4b96-ad58-4649ae578ca4'

        self.ONSM = ['HAY ALMANAR RTE AHFIR', 'HAY JARF LAKHDAR', 'HAY AMAL 1 et 2', 'LOTS SALAMA', 'HAY ANGADI',
                'hay ALOUAHDA RTE TAZA', 'HAY ENNAHDA ( ex amr alboulissi)', 'hay khaloufi , hay alouafae',
                'LOTS AZIZI , AZOUHOUr , BENKHALDOUN RTE JERADA', 'HAY ALJAWHARA , BENAZZI', 'LOTS MILITAIRE OLM',
                'HAY LAMHALLA RHA TBOUL', 'HAY ESSALAM AOUINAT ESSARAK', 'HAY ASSAADA AOUINAT ESSARAK',
                "HAY STADE D'HONNEUR", 'HAY BENKHIRANE', 'LOT IRIS', 'HAY ALHIKMA , HAY ALAIRFANE', 'HAY BELAMRAH',
                'HAY BOUARFA', 'HAY ERRABIE , LOTS BELLAOUI', 'HAY ALQODS', 'HAY ALANDALOUS', 'HAY HAKKOU',
                'ADOUHA']
        self.DK = ['CGI Rte Ahfir', 'LOTS TALHAOUI ET MIR', 'RIAD ISLY', 'LOTS ALJ RTE AHFIR',
              'CITE GOLF RTE AHFIR', 'Sadrat bouamoud hay med belkhdar', 'HAY JARF LAKHDAR',
              'HAY ALHASSANI ( KOULOUCH)', 'TOBA INTERIEUR', 'HAY TANMIA', 'HAY MAURITANIA', 'Hay tenis',
              'HAY ALBOUSTANE', 'HAY ANNASR , RJAF ALLAH , HAY GHAR ALBAROUD']
        self.SY = ['HAY ALFATH , LOTS ANGAD , HAY ESSALAM , DEM', 'HAY MY MILOUD',
              'HAY LABYAYED , HAY FARAH , HAY MEXICO , LOTS MINICIPALITE', 'HAY LAHBOUS DEM', 'HAY MY MUSTAPHA',
              'HAY EZZAYTOUNE', 'HAY AGDAL', 'HAY ANAJD , HASSANIA']
        self.SZ = ['CENTRE VILLE', 'HAY BOUDIR', 'DERB MBASSOU', 'HAY RAS OUSFOUR', 'LOTS BOUKNADEL RTE TAIRET',
              'HAY AL MASSIRA', 'HAY LAMHALLA , BD ALHIJAZ', 'CENTRE VILLE (VILLE NOUVELLE)', 'MEDINA']


        self.allhoods = self.ONSM + self.DK + self.SY + self.SZ

        #print(f'the data that i want ====>  {self.data}')
        #TODO : The Data ===> ['19', 'RIME ZAHIRI', 'F652030', 'HAY SAADA RUE AL BATRIQ', 'TDE', '08-17', 'None']
        self.strr.addItems(self.allhoods)
        # print(self.allhoods[self.allhoods.index(self.data[7]) + 1])
        if self.data:
            self.matrr.setText(data[0])
            self.Fname.setText(str(self.data[1]).split(' ')[0])
            self.Lname.setText(str(self.data[1]).split(' ')[1])
            self.CIN.setText(self.data[2])
            self.ad.setText(self.data[3])
            self.grps.setCurrentIndex(grps_names.index(self.data[4]))

            self.strr.setCurrentIndex(self.allhoods.index(self.data[7]) + 1 if self.data[7] and self.data[7] != 'None' else 0 )

        self.setWindowTitle('Add new Agent')

        self.back_icon.installEventFilter(self)
        self.back_icon.setPixmap(QtGui.QPixmap('src/img/back.png'))
        self.back_icon.setScaledContents(True)

        #print('+' *100)
        #print(self.allhoods)
        cnx.close()
        self.pic.installEventFilter(self)
        self.refreshPic()

    def refreshPic(self):
        image = QtGui.QImage()
        image.loadFromData(requests.get(self.agent_pic_path).content)
        self.pic.setPixmap(QtGui.QPixmap(image))
        self.pic.setScaledContents(True)

    def eventFilter(self, s, e):
        if s is self.back_icon:
            if e.type() == QtCore.QEvent.MouseButtonPress:
                self.close()
            if e.type() == QtCore.QEvent.Enter:
                self.back_icon.setStyleSheet("border: 1px solid black; border-radius: 25px;")
                #print('mouse in ')
            elif e.type() == QtCore.QEvent.Leave:
                self.back_icon.setStyleSheet('border-color: 0px;')
                #print('mouse out')
        if e.type() == QtCore.QEvent.MouseButtonPress and s is self.pic:
            filename, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Open Picture", "~", "Sound Files (*.png *.jpg *.jpeg )")
            if ntpath.isfile(filename):
                self.agent_pic_path = uploadPicToFireBase(filename)
                self.refreshPic()
                # if self.action == "edit":
                #     cnx = con()
                #     cur = cnx.cursor()
                #
                #
                #     cnx.close()


        return super(AddAgent, self).eventFilter(s, e)

    def closeEvent(self, event):
        self.agentPage = Agents(self.role)
        self.agentPage.show()
        self.close()



    def save(self):

        if len(self.Fname.text()) > 3 and len(self.Lname.text()) > 2 and len(self.CIN.text()) > 3 and len(self.ad.text()) > 2 and self.strr.currentIndex() != 0 and len(self.matrr.text()) > 1:
            cnx = con()
            cur = cnx.cursor()
            # cur.execute(f'''select concat(firstName , " ", LastName) from agents where matrr like "{self.matrr.text()}"''')
            # if agent:=cur.fetchone()[0]:
            #     notif(self, title="Agent Allready exists", msg=f"This Matriculation alreaddy exists on the database for Agent : {agent}")
            #     return

            cur.execute(f'select id from grps where name like "{self.grps.currentText()}"')
            grp = int(cur.fetchone()[0])
            zone = ''



            if self.strr.currentText() in self.ONSM:
                zone = "Oued Nachef Sidi Maafa"
            elif self.strr.currentText() in self.DK:
                zone = "Sidi Driss EL Kadi"
            elif self.strr.currentText() in self.SY:
                zone = "Sidi Yahya"
            elif  self.strr.currentText() in self.SZ:
                zone = "Sidi Ziane"

            if self.action == 'save':
                cur.execute(
                    f'select concat(firstName , " ", LastName) from agents where matrr like "{self.matrr.text()}" ')
                a = cur.fetchone()
                print(a)
                print(f'type of a : {type(a)}')
                if  a is not None:
                    #print('thsi Agent alredy exists ')
                    self.Fname.setStyleSheet('border : 2px solid red')
                    self.Lname.setStyleSheet('border : 2px solid red')
                    notif(self, title='Agent Already exists', msg=f'This Matriculation already exists in the Database : - {a}')
                    return
                else:
                    cur.execute(f'insert into agents (matrr, firstName, LastName, CIN, address, grp, zone, street, pic) value ("{self.matrr.text()}", "{self.Fname.text()}", "{self.Lname.text()}", "{self.CIN.text()}", "{self.ad.text()}", {grp}, "{zone}", "{self.strr.currentText()}", "{self.agent_pic_path}");')
                    cnx.commit()
            elif self.action == 'edit':
                    self.matrr.setEnabled(False)
                    query = f'''
                            update agents set  
                                            firstName = "{self.Fname.text()}",
                                            LastName = "{self.Lname.text()}",
                                            CIN = "{self.CIN.text()}",
                                            address = "{self.ad.text()}",
                                            grp = {grp},
                                            zone = "{zone}",
                                            street = "{self.strr.currentText()}",
                                            pic = "{self.agent_pic_path}"
                                            where matrr = "{self.data[0]}"
                        '''
                    print(query)
                    cur.execute(query=query)
                    cnx.commit()
                    # self.agents = Agents(self.role)
                    # self.agents.show()
                    # self.close()
            self.Fname.clear()
            self.Lname.clear()
            self.CIN.clear()
            self.ad.clear()
            self.grps.setCurrentIndex(0)
            self.strr.setCurrentIndex(0)
            self.agent_pic_path = 'https://firebasestorage.googleapis.com/v0/b/saccomxd-stm-yassine-baghdadi.appspot.com/o/profile.png?alt=media&token=30aaedee-2f7d-4b96-ad58-4649ae578ca4'
            self.refreshPic()


            cnx.close()

        else:
            notif(self, title="Error ", msg='you have to write all the infos')



if __name__ == '__main__' :
    app = QtWidgets.QApplication(sys.argv)
    splash_wn = Splash()
    splash_wn.show()
    sys.exit(app.exec_())